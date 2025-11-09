"""
Class-based views for inventory management.
"""
import uuid
import json
import logging
import traceback
from decimal import Decimal
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import JsonResponse, HttpResponseBadRequest
from django.db import transaction
from django.db.models import Sum, F, Q
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.contrib import messages

from .models import Item, Batch, TransactionLog, Order, OrderItem, Allocation, Shipment, Notification, Return
from .forms import ItemForm, BatchForm, PickForm, PackForm, ShipForm, ReturnForm, ReturnProcessForm, BulkImportForm, OrderForm, OrderItemInlineFormSet
from .services.allocation import allocate_order, AllocationError, OrderNotFoundError
from .services.batch_processor import process_order_queue_batch
from .services.notifications import (
    send_shipment_notification,
    send_low_stock_alert,
    webhook_shipment_created,
    webhook_order_fulfilled,
)
from .services.undo_redo import perform_undo, perform_redo, push_undo_operation
from .services.notifications_helper import (
    notify,
    get_unread_count,
    get_recent_notifications,
    mark_as_read,
    mark_all_as_read,
)
from .integrations import validate_webhook_signature

logger = logging.getLogger(__name__)




class HtmxResponseMixin:
    """Mixin to return partial templates for htmx requests."""

    htmx_template_name = None

    def get_template_names(self):
        if self.request.headers.get("HX-Request") and self.htmx_template_name:
            return [self.htmx_template_name]
        return super().get_template_names()


# =============================
# Dashboard View
# =============================

class DashboardView(TemplateView):
    """Main dashboard showing system overview."""
    template_name = "inventory/dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Basic stats
        context['total_items'] = Item.objects.count()
        context['total_batches'] = Batch.objects.count()
        context['total_orders'] = Order.objects.count()
        context['pending_orders'] = Order.objects.filter(status=Order.STATUS_NEW).count()
        context['allocated_orders'] = Order.objects.filter(status=Order.STATUS_ALLOCATED).count()
        
        # Recent orders
        context['recent_orders'] = Order.objects.select_related().all()[:5]
        
        # Low stock items
        context['low_stock_items'] = Item.objects.filter(
            reorder_threshold__gt=0
        ).annotate(
            current_qty=Sum('batches__available_qty')
        ).filter(
            current_qty__lte=F('reorder_threshold')
        )[:5]
        
        # Expiring soon (next 30 days)
        from datetime import timedelta
        expiry_threshold = timezone.now().date() + timedelta(days=30)
        context['expiring_batches'] = Batch.objects.filter(
            expiry_date__lte=expiry_threshold,
            expiry_date__gte=timezone.now().date(),
            status=Batch.STATUS_AVAILABLE
        ).select_related('item').order_by('expiry_date')[:5]
        
        return context


# =============================
# Item Views
# =============================

class ItemListView(HtmxResponseMixin, ListView):
    model = Item
    template_name = "inventory/item_list.html"
    htmx_template_name = "inventory/partials/item_table.html"
    context_object_name = "items"
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.GET.get("search", "")
        if search:
            qs = qs.filter(sku__icontains=search) | qs.filter(name__icontains=search)
        return qs


class ItemDetailView(HtmxResponseMixin, DetailView):
    model = Item
    template_name = "inventory/item_detail.html"
    htmx_template_name = "inventory/partials/item_detail_content.html"
    context_object_name = "item"
    slug_field = "sku"
    slug_url_kwarg = "sku"


class ItemCreateView(CreateView):
    model = Item
    form_class = ItemForm
    template_name = "inventory/item_form.html"
    success_url = reverse_lazy("inventory:item-list")


class ItemUpdateView(UpdateView):
    model = Item
    form_class = ItemForm
    template_name = "inventory/item_form.html"
    slug_field = "sku"
    slug_url_kwarg = "sku"

    def get_success_url(self):
        return reverse_lazy("inventory:item-detail", kwargs={"sku": self.object.sku})


class ItemDeleteView(DeleteView):
    model = Item
    template_name = "inventory/item_confirm_delete.html"
    success_url = reverse_lazy("inventory:item-list")
    slug_field = "sku"
    slug_url_kwarg = "sku"


# =============================
# Batch Views
# =============================

class BatchListView(HtmxResponseMixin, ListView):
    model = Batch
    template_name = "inventory/batch_list.html"
    htmx_template_name = "inventory/partials/batch_table.html"
    context_object_name = "batches"
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().select_related("item")
        search = self.request.GET.get("search", "")
        if search:
            qs = qs.filter(lot_no__icontains=search) | qs.filter(item__sku__icontains=search)
        return qs


class BatchDetailView(DetailView):
    model = Batch
    template_name = "inventory/batch_detail.html"
    context_object_name = "batch"


class BatchCreateView(CreateView):
    model = Batch
    form_class = BatchForm
    template_name = "inventory/batch_form.html"
    success_url = reverse_lazy("inventory:batch-list")


class BatchUpdateView(UpdateView):
    model = Batch
    form_class = BatchForm
    template_name = "inventory/batch_form.html"
    success_url = reverse_lazy("inventory:batch-list")


class BatchDeleteView(DeleteView):
    model = Batch
    template_name = "inventory/batch_confirm_delete.html"
    success_url = reverse_lazy("inventory:batch-list")


# =============================
# Receive View
# =============================

class ReceiveView(TemplateView):
    """View for receiving multiple batches for a single item."""
    template_name = "inventory/receive.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["items"] = Item.objects.all().order_by("sku")
        return context

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action")
        
        if action == "preview":
            return self.preview(request)
        elif action == "commit":
            return self.commit(request)
        
        return JsonResponse({"error": "Invalid action"}, status=400)

    def preview(self, request):
        """Validate batch data and return preview for htmx."""
        item_id = request.POST.get("item_id")
        if not item_id:
            return render(request, "inventory/partials/receive_preview.html", {
                "errors": ["Item is required."]
            })

        try:
            item = Item.objects.get(pk=item_id)
        except Item.DoesNotExist:
            return render(request, "inventory/partials/receive_preview.html", {
                "errors": ["Item not found."]
            })

        # Parse batch rows from POST data
        batches_data = []
        errors = []
        row_index = 0
        
        while True:
            lot_no = request.POST.get(f"batch_{row_index}_lot_no")
            if lot_no is None:
                break
            
            qty_str = request.POST.get(f"batch_{row_index}_qty", "0")
            expiry_str = request.POST.get(f"batch_{row_index}_expiry", "")
            
            # Validation
            row_errors = []
            if not lot_no.strip():
                row_errors.append(f"Row {row_index + 1}: Lot number is required")
            else:
                # Check if lot already exists
                if Batch.objects.filter(item=item, lot_no=lot_no).exists():
                    row_errors.append(f"Row {row_index + 1}: Lot '{lot_no}' already exists for this item")
            
            try:
                qty = Decimal(qty_str)
                if qty <= 0:
                    row_errors.append(f"Row {row_index + 1}: Quantity must be positive")
            except (ValueError, TypeError):
                row_errors.append(f"Row {row_index + 1}: Invalid quantity")
                qty = Decimal("0")
            
            if row_errors:
                errors.extend(row_errors)
            else:
                batches_data.append({
                    "lot_no": lot_no,
                    "qty": qty,
                    "expiry": expiry_str if expiry_str else None,
                })
            
            row_index += 1

        if not batches_data and not errors:
            errors.append("At least one batch row is required.")

        context = {
            "item": item,
            "batches_data": batches_data,
            "errors": errors,
            "can_commit": not errors and batches_data,
        }
        
        return render(request, "inventory/partials/receive_preview.html", context)

    def commit(self, request):
        """Create batches and transaction logs."""
        item_id = request.POST.get("item_id")
        if not item_id:
            return JsonResponse({"error": "Item is required"}, status=400)

        try:
            item = Item.objects.get(pk=item_id)
        except Item.DoesNotExist:
            return JsonResponse({"error": "Item not found"}, status=400)

        # Parse and validate batch data
        batches_to_create = []
        row_index = 0
        
        while True:
            lot_no = request.POST.get(f"batch_{row_index}_lot_no")
            if lot_no is None:
                break
            
            qty_str = request.POST.get(f"batch_{row_index}_qty", "0")
            expiry_str = request.POST.get(f"batch_{row_index}_expiry", "")
            
            if not lot_no.strip():
                row_index += 1
                continue
            
            try:
                qty = Decimal(qty_str)
                if qty <= 0:
                    row_index += 1
                    continue
            except (ValueError, TypeError):
                row_index += 1
                continue
            
            # Check duplicate
            if Batch.objects.filter(item=item, lot_no=lot_no).exists():
                row_index += 1
                continue
            
            batches_to_create.append({
                "lot_no": lot_no,
                "qty": qty,
                "expiry": expiry_str if expiry_str else None,
            })
            
            row_index += 1

        if not batches_to_create:
            return JsonResponse({"error": "No valid batches to create"}, status=400)

        # Create batches and logs in a transaction
        created_count = 0
        with transaction.atomic():
            for batch_data in batches_to_create:
                batch = Batch.objects.create(
                    item=item,
                    lot_no=batch_data["lot_no"],
                    received_qty=batch_data["qty"],
                    available_qty=batch_data["qty"],
                    expiry_date=batch_data["expiry"] or None,
                    status=Batch.STATUS_AVAILABLE,
                )
                
                # Create transaction log
                TransactionLog.objects.create(
                    user=request.user if request.user.is_authenticated else None,
                    type=TransactionLog.TYPE_RECEIPT,
                    qty=batch_data["qty"],
                    item=item,
                    batch=batch,
                    meta={"lot_no": batch_data["lot_no"]},
                )
                created_count += 1

        # Return success response for htmx
        return render(request, "inventory/partials/receive_success.html", {
            "item": item,
            "created_count": created_count,
        })


# =============================
# Order Allocation View
# =============================

class AllocateOrderView(UserPassesTestMixin, View):
    """View to trigger allocation for an order (manager-only)."""

    def test_func(self):
        # Only staff/superusers can allocate
        return self.request.user.is_staff or self.request.user.is_superuser

    def get(self, request, *args, **kwargs):
        """Show allocation confirmation page."""
        order_id = kwargs.get("order_id")
        order = get_object_or_404(Order.objects.prefetch_related('items__item'), pk=order_id)
        
        context = {
            "order": order,
            "order_items": order.items.select_related('item').all(),
        }
        return render(request, "inventory/allocate_confirm.html", context)

    def post(self, request, *args, **kwargs):
        order_id = kwargs.get("order_id")
        
        try:
            result = allocate_order(order_id, user=request.user)
            
            # Fetch the order for redirect
            order = Order.objects.get(pk=order_id)
            
            # Add success message
            messages.success(
                request, 
                f"Successfully allocated stock for order {order.order_no}. "
                f"Allocated {result.get('items_allocated', 0)} items."
            )
            
            # Redirect to order detail page
            return redirect('inventory:order-detail', pk=order_id)
            
        except OrderNotFoundError as e:
            messages.error(request, f"Order not found: {str(e)}")
            return redirect('inventory:order-list')
            
        except AllocationError as e:
            messages.error(request, f"Allocation failed: {str(e)}")
            return redirect('inventory:order-detail', pk=order_id)
            
        except Exception as e:
            messages.error(request, f"Unexpected error: {str(e)}")
            return redirect('inventory:order-detail', pk=order_id)


# =============================
# Deallocate Order View
# =============================

class DeallocateOrderView(UserPassesTestMixin, View):
    """View to deallocate/release stock from an order (manager-only)."""

    def test_func(self):
        # Only staff/superusers can deallocate
        return self.request.user.is_staff or self.request.user.is_superuser

    def get(self, request, *args, **kwargs):
        """Show deallocation confirmation page."""
        order_id = kwargs.get("order_id")
        order = get_object_or_404(
            Order.objects.prefetch_related('items__item', 'items__allocations__batch'), 
            pk=order_id
        )
        
        # Get all allocations for this order
        allocations = []
        for order_item in order.items.all():
            for allocation in order_item.allocations.all():
                allocations.append({
                    'allocation': allocation,
                    'order_item': order_item,
                    'item': order_item.item,
                    'batch': allocation.batch,
                })
        
        context = {
            "order": order,
            "allocations": allocations,
        }
        return render(request, "inventory/deallocate_confirm.html", context)

    def post(self, request, *args, **kwargs):
        """Perform deallocation - release stock back to batches."""
        order_id = kwargs.get("order_id")
        order = get_object_or_404(Order, pk=order_id)
        
        try:
            with transaction.atomic():
                total_deallocated = 0
                allocation_count = 0
                
                # Get all allocations for this order
                for order_item in order.items.select_for_update().all():
                    for allocation in order_item.allocations.select_for_update().all():
                        # Release stock back to batch
                        batch = allocation.batch
                        batch.available_qty += allocation.qty_allocated
                        batch.save()
                        
                        # Log the deallocation
                        TransactionLog.objects.create(
                            batch=batch,
                            item=order_item.item,
                            order=order,
                            type=TransactionLog.TYPE_DEALLOCATE,
                            qty=allocation.qty_allocated,
                            user=request.user,
                            meta={
                                "reason": "manual_deallocation",
                                "allocation_id": allocation.id,
                                "order_no": order.order_no,
                            }
                        )
                        
                        total_deallocated += allocation.qty_allocated
                        allocation_count += 1
                        
                        # Delete the allocation
                        allocation.delete()
                    
                    # Reset order item allocated quantity
                    order_item.qty_allocated = Decimal("0")
                    order_item.save()
                
                # Reset order status to NEW
                order.status = Order.STATUS_NEW
                order.save()
                
                # Create notification
                notify(
                    title=f"Order {order.order_no} deallocated",
                    message=f"Released {total_deallocated} items from {allocation_count} allocations back to inventory.",
                    level=Notification.LEVEL_INFO,
                    user=request.user,
                )
                
                # Add success message
                from django.contrib import messages
                messages.success(
                    request, 
                    f"Successfully deallocated order {order.order_no}. "
                    f"Released {total_deallocated} items from {allocation_count} allocations back to inventory."
                )
                
                # Redirect to order detail page
                from django.shortcuts import redirect
                return redirect('inventory:order-detail', pk=order_id)
                
        except Exception as e:
            from django.contrib import messages
            messages.error(request, f"Deallocation failed: {str(e)}")
            from django.shortcuts import redirect
            return redirect('inventory:order-detail', pk=order_id)


# =============================
# Pick View
# =============================

class PickView(UserPassesTestMixin, View):
    """View for confirming picked quantities against allocations."""
    
    def test_func(self):
        return self.request.user.is_authenticated
    
    def get(self, request, *args, **kwargs):
        """Display pick list for an order."""
        order_id = kwargs.get("order_id")
        order = get_object_or_404(
            Order.objects.prefetch_related(
                "items__allocations__batch__item",
                "items__item"
            ),
            pk=order_id
        )
        
        # Build list of allocations to pick
        pick_items = []
        for order_item in order.items.all():
            for allocation in order_item.allocations.all():
                pick_items.append({
                    "allocation": allocation,
                    "order_item": order_item,
                    "item": order_item.item,
                    "batch": allocation.batch,
                })
        
        context = {
            "order": order,
            "pick_items": pick_items,
        }
        return render(request, "inventory/pick.html", context)
    
    def post(self, request, *args, **kwargs):
        """Process picked quantities."""
        order_id = kwargs.get("order_id")
        order = get_object_or_404(Order, pk=order_id)
        
        try:
            with transaction.atomic():
                pick_results = []
                
                # Process each allocation's picked quantity
                for key, value in request.POST.items():
                    if key.startswith("qty_picked_"):
                        allocation_id = int(key.split("_")[-1])
                        qty_picked = Decimal(value)
                        
                        allocation = Allocation.objects.select_for_update().get(pk=allocation_id)
                        
                        # Update allocation with picked quantity
                        if qty_picked != allocation.qty_allocated:
                            # Log discrepancy
                            TransactionLog.objects.create(
                                batch=allocation.batch,
                                item=allocation.order_item.item,
                                order=allocation.order_item.order,
                                qty=-qty_picked,
                                type=TransactionLog.TYPE_ADJUST,
                                user=request.user,
                                meta={
                                    "reason": "pick_adjust",
                                    "order_item_id": allocation.order_item.id,
                                    "allocated": str(allocation.qty_allocated),
                                    "picked": str(qty_picked),
                                },
                            )
                        
                        pick_results.append({
                            "allocation": allocation,
                            "qty_picked": qty_picked,
                            "qty_allocated": allocation.qty_allocated,
                            "item_sku": allocation.order_item.item.sku,
                        })
                
                # Update order status to PICKED
                order.status = Order.STATUS_PICKED
                order.save()
                
                # Add success message
                from django.contrib import messages
                messages.success(
                    request, 
                    f"Successfully picked order {order.order_no}. Ready for packing."
                )
                
                # Redirect to order detail or pack view
                from django.shortcuts import redirect
                return redirect('inventory:order-detail', pk=order_id)
                
        except Exception as e:
            context = {
                "success": False,
                "error_message": str(e),
            }
            return render(request, "inventory/partials/pick_result.html", context)


# =============================
# Pack View
# =============================

class PackView(UserPassesTestMixin, View):
    """View for confirming packing and optionally adjusting quantities."""
    
    def test_func(self):
        return self.request.user.is_authenticated
    
    def get(self, request, *args, **kwargs):
        """Display pack confirmation form."""
        order_id = kwargs.get("order_id")
        order = get_object_or_404(
            Order.objects.prefetch_related(
                "items__allocations__batch",
                "items__item"
            ),
            pk=order_id
        )
        
        # Calculate total picked/allocated per order item
        pack_items = []
        for order_item in order.items.all():
            total_allocated = sum(
                a.qty_allocated for a in order_item.allocations.all()
            )
            pack_items.append({
                "order_item": order_item,
                "total_allocated": total_allocated,
            })
        
        context = {
            "order": order,
            "pack_items": pack_items,
        }
        return render(request, "inventory/pack.html", context)
    
    def post(self, request, *args, **kwargs):
        """Process packed quantities."""
        order_id = kwargs.get("order_id")
        order = get_object_or_404(Order, pk=order_id)
        
        try:
            with transaction.atomic():
                pack_results = []
                
                # Process each order item's packed quantity
                for key, value in request.POST.items():
                    if key.startswith("qty_packed_"):
                        order_item_id = int(key.split("_")[-1])
                        qty_packed = Decimal(value)
                        
                        order_item = OrderItem.objects.select_for_update().get(pk=order_item_id)
                        total_allocated = sum(
                            a.qty_allocated for a in order_item.allocations.all()
                        )
                        
                        # Update order item with packed quantity
                        order_item.qty_picked = qty_packed
                        order_item.save()
                        
                        if qty_packed != total_allocated:
                            # Log packing adjustment
                            notes_key = f"notes_{order_item_id}"
                            notes = request.POST.get(notes_key, "Packing adjustment")
                            
                            TransactionLog.objects.create(
                                item=order_item.item,
                                order=order_item.order,
                                qty=-qty_packed,
                                type=TransactionLog.TYPE_ADJUST,
                                user=request.user,
                                meta={
                                    "reason": "pack_adjust",
                                    "order_item_id": order_item_id,
                                    "allocated": str(total_allocated),
                                    "packed": str(qty_packed),
                                    "notes": notes,
                                },
                            )
                        
                        pack_results.append({
                            "order_item": order_item,
                            "qty_packed": qty_packed,
                            "total_allocated": total_allocated,
                        })

                # Update order status to PACKED after packing
                order.status = Order.STATUS_PACKED
                order.save(update_fields=["status"])
                
                # Add success message
                from django.contrib import messages
                messages.success(
                    request, 
                    f"Successfully packed order {order.order_no}. Ready for shipment."
                )
                
                # Redirect to order detail to show ship button
                from django.shortcuts import redirect
                return redirect('inventory:order-detail', pk=order_id)
                
        except Exception as e:
            context = {
                "success": False,
                "error_message": str(e),
            }
            return render(request, "inventory/partials/pack_result.html", context)


# =============================
# Ship View
# =============================

class ShipView(UserPassesTestMixin, View):
    """View for creating shipment and finalizing order."""
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser
    
    def get(self, request, *args, **kwargs):
        """Display shipment form."""
        order_id = kwargs.get("order_id")
        order = get_object_or_404(
            Order.objects.prefetch_related(
                "items__allocations__batch",
                "items__item"
            ),
            pk=order_id
        )
        
        # Check if order is ready to ship (must be allocated or picked)
        if order.status not in [Order.STATUS_ALLOCATED, Order.STATUS_PICKED]:
            from django.contrib import messages
            messages.warning(request, f"Order must be allocated or picked before shipping. Current status: {order.get_status_display()}")
            from django.shortcuts import redirect
            return redirect('inventory:order-detail', pk=order_id)
        
        context = {
            "order": order,
        }
        return render(request, "inventory/ship.html", context)
    
    def post(self, request, *args, **kwargs):
        """Create shipment and finalize order."""
        order_id = kwargs.get("order_id")
        order = get_object_or_404(Order, pk=order_id)
        
        try:
            with transaction.atomic():
                # Generate unique tracking number and shipment number
                tracking_no = uuid.uuid4()
                shipment_no = f"SHIP-{order.order_no}-{timezone.now().strftime('%Y%m%d%H%M%S')}"
                
                # Extract form data
                carrier = request.POST.get("carrier", "")
                shipping_address = request.POST.get("shipping_address", "")
                notes = request.POST.get("notes", "")
                
                # Create shipment record
                shipment = Shipment.objects.create(
                    order=order,
                    shipment_no=shipment_no,
                    tracking_no=tracking_no,
                    carrier=carrier,
                    shipping_address=shipping_address,
                    shipped_at=timezone.now(),
                    notes=notes,
                )
                
                # Finalize consumption: Log shipment transactions and clean up allocations
                # Note: Batch quantities were already reduced during allocation,
                # so we don't reduce them again here. We just log the shipment and delete allocations.
                for order_item in order.items.all():
                    for allocation in order_item.allocations.all():
                        # Log consumption transaction (quantity already deducted during allocation)
                        TransactionLog.objects.create(
                            batch=allocation.batch,
                            item=order_item.item,
                            order=order,
                            shipment=shipment,
                            qty=-allocation.qty_allocated,
                            type=TransactionLog.TYPE_SHIP,
                            user=request.user,
                            meta={
                                "order_item_id": order_item.id,
                                "carrier": carrier or "unknown",
                                "tracking_no": str(tracking_no),
                                "notes": notes,
                            },
                        )
                        
                        # Check for low stock
                        item = order_item.item
                        total_qty = item.total_quantity()
                        if item.reorder_threshold and total_qty <= item.reorder_threshold:
                            send_low_stock_alert(item, total_qty)
                
                # Delete allocations after shipping (stock already deducted during allocation)
                Allocation.objects.filter(order_item__order=order).delete()
                
                # Update order status to SHIPPED
                order.status = Order.STATUS_SHIPPED
                order.save()
                
                # Create notification
                Notification.objects.create(
                    user=request.user,
                    message=f"Order {order.order_no} shipped with tracking {tracking_no}",
                    level=Notification.LEVEL_INFO,
                )
                
                # Send email notification
                send_shipment_notification(shipment)
                
                # Trigger webhook
                webhook_shipment_created(shipment)
                
                # If all items fulfilled, trigger order fulfilled webhook
                if order.status == Order.STATUS_SHIPPED:
                    webhook_order_fulfilled(order)
                
                # Add success message
                from django.contrib import messages
                messages.success(
                    request, 
                    f"Order {order.order_no} successfully shipped! Tracking: {tracking_no}"
                )
                
                # Redirect to order detail
                from django.shortcuts import redirect
                return redirect('inventory:order-detail', pk=order_id)
                
        except Exception as e:
            from django.contrib import messages
            messages.error(request, f"Shipping failed: {str(e)}")
            from django.shortcuts import redirect
            return redirect('inventory:order-ship', order_id=order_id)


# =============================
# Deliver View
# =============================

class DeliverView(UserPassesTestMixin, View):
    """View for marking an order as delivered."""
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser
    
    def get(self, request, *args, **kwargs):
        """Display delivery confirmation form."""
        order_id = kwargs.get("order_id")
        order = get_object_or_404(Order, pk=order_id)
        
        # Check if order is shipped
        if order.status != Order.STATUS_SHIPPED:
            from django.contrib import messages
            messages.warning(request, "Order must be shipped before marking as delivered.")
            from django.shortcuts import redirect
            return redirect('inventory:order-detail', pk=order_id)
        
        # Get shipment details
        shipment = order.shipments.first()
        
        context = {
            "order": order,
            "shipment": shipment,
        }
        return render(request, "inventory/deliver.html", context)
    
    def post(self, request, *args, **kwargs):
        """Mark order as delivered."""
        order_id = kwargs.get("order_id")
        order = get_object_or_404(Order, pk=order_id)
        
        try:
            with transaction.atomic():
                # Update order status to DELIVERED
                order.status = Order.STATUS_DELIVERED
                order.save()
                
                # Update shipment status if exists
                shipment = order.shipments.first()
                if shipment:
                    shipment.status = Shipment.STATUS_DELIVERED
                    shipment.delivered_at = timezone.now()
                    shipment.save()
                
                # Create notification
                Notification.objects.create(
                    user=request.user,
                    message=f"Order {order.order_no} has been delivered to {order.customer_name}",
                    level=Notification.LEVEL_INFO,
                )
                
                # Add success message
                from django.contrib import messages
                messages.success(
                    request, 
                    f"Order {order.order_no} marked as delivered!"
                )
                
                # Redirect to order detail
                from django.shortcuts import redirect
                return redirect('inventory:order-detail', pk=order_id)
                
        except Exception as e:
            from django.contrib import messages
            messages.error(request, f"Failed to mark as delivered: {str(e)}")
            from django.shortcuts import redirect
            return redirect('inventory:order-deliver', order_id=order_id)


# =============================
# RMA (Return) Views
# =============================

class CreateReturnView(LoginRequiredMixin, CreateView):
    """View for creating a return (RMA)."""
    
    model = Return
    form_class = ReturnForm
    template_name = "inventory/return_create.html"
    success_url = reverse_lazy("inventory:return-list")
    
    def form_valid(self, form):
        # Generate unique return number
        return_no = f"RMA-{uuid.uuid4().hex[:8].upper()}"
        form.instance.return_no = return_no
        return super().form_valid(form)


class ReturnListView(LoginRequiredMixin, ListView):
    """View for listing all returns."""
    
    model = Return
    template_name = "inventory/return_list.html"
    context_object_name = "returns"
    paginate_by = 20
    
    def get_queryset(self):
        qs = super().get_queryset().select_related("order_item__item", "order_item__order")
        status = self.request.GET.get("status", "")
        if status:
            qs = qs.filter(status=status)
        return qs


class ReturnDetailView(LoginRequiredMixin, DetailView):
    """View for viewing return details."""
    
    model = Return
    template_name = "inventory/return_detail.html"
    context_object_name = "return_obj"
    
    def get_queryset(self):
        return super().get_queryset().select_related(
            "order_item__item",
            "order_item__order"
        )


class ProcessReturnView(UserPassesTestMixin, View):
    """View for processing a return with disposition options."""
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser
    
    def get(self, request, *args, **kwargs):
        """Display return processing form."""
        return_id = kwargs.get("return_id")
        return_obj = get_object_or_404(
            Return.objects.select_related("order_item__item"),
            pk=return_id
        )
        
        form = ReturnProcessForm(initial={
            "return_id": return_id,
            "qty_accepted": return_obj.qty_returned,
        })
        
        context = {
            "return_obj": return_obj,
            "form": form,
        }
        return render(request, "inventory/process_return.html", context)
    
    def post(self, request, *args, **kwargs):
        """Process return based on disposition."""
        return_id = kwargs.get("return_id")
        return_obj = get_object_or_404(Return, pk=return_id)
        
        form = ReturnProcessForm(request.POST)
        
        if not form.is_valid():
            context = {
                "return_obj": return_obj,
                "form": form,
            }
            return render(request, "inventory/process_return.html", context)
        
        disposition = form.cleaned_data["disposition"]
        qty_accepted = form.cleaned_data["qty_accepted"]
        notes = form.cleaned_data.get("notes", "")
        
        try:
            with transaction.atomic():
                order_item = return_obj.order_item
                item = order_item.item
                
                # Find original allocations to determine source batches
                original_allocations = order_item.allocations.all()
                
                if disposition == ReturnProcessForm.DISPOSITION_RESTOCK_ORIGINAL:
                    # Restock to original batch(es)
                    if original_allocations.exists():
                        # Restock proportionally to original allocations
                        allocation = original_allocations.first()
                        batch = Batch.objects.select_for_update().get(pk=allocation.batch.pk)
                        batch.available_qty += qty_accepted
                        batch.save()
                        
                        # Log transaction
                        TransactionLog.objects.create(
                            batch=batch,
                            item=item,
                            order=order_item.order,
                            qty=qty_accepted,
                            type=TransactionLog.TYPE_ADJUST,
                            user=request.user,
                            meta={
                                "reason": "return_restock",
                                "return_no": return_obj.return_no,
                                "order_item_id": order_item.id,
                                "notes": notes,
                            },
                        )
                        
                        return_obj.status = Return.STATUS_RESTOCKED
                    else:
                        raise ValueError("No original allocations found for restock.")
                
                elif disposition == ReturnProcessForm.DISPOSITION_RESTOCK_NEW:
                    # Create new return batch with special lot number
                    return_lot_no = f"RETURN-{uuid.uuid4().hex[:8].upper()}"
                    
                    batch = Batch.objects.create(
                        item=item,
                        lot_no=return_lot_no,
                        received_qty=qty_accepted,
                        available_qty=qty_accepted,
                        status=Batch.STATUS_AVAILABLE,
                    )
                    
                    # Log transaction
                    TransactionLog.objects.create(
                        batch=batch,
                        item=item,
                        order=order_item.order,
                        qty=qty_accepted,
                        type=TransactionLog.TYPE_RECEIPT,
                        user=request.user,
                        meta={
                            "reason": "return_new_batch",
                            "return_no": return_obj.return_no,
                            "order_item_id": order_item.id,
                            "notes": notes,
                        },
                    )
                    
                    return_obj.status = Return.STATUS_RESTOCKED
                
                elif disposition == ReturnProcessForm.DISPOSITION_QUARANTINE:
                    # Create quarantine batch
                    quarantine_lot_no = f"QUARANTINE-{uuid.uuid4().hex[:8].upper()}"
                    
                    batch = Batch.objects.create(
                        item=item,
                        lot_no=quarantine_lot_no,
                        received_qty=qty_accepted,
                        available_qty=Decimal("0"),  # Not available for allocation
                        status=Batch.STATUS_QUARANTINE,
                    )
                    
                    # Log transaction
                    TransactionLog.objects.create(
                        batch=batch,
                        item=item,
                        order=order_item.order,
                        qty=qty_accepted,
                        type=TransactionLog.TYPE_RECEIPT,
                        user=request.user,
                        meta={
                            "reason": "return_quarantine",
                            "return_no": return_obj.return_no,
                            "order_item_id": order_item.id,
                            "notes": notes,
                        },
                    )
                    
                    return_obj.status = Return.STATUS_QUARANTINED
                
                elif disposition == ReturnProcessForm.DISPOSITION_SCRAP:
                    # Just mark as scrapped, no inventory adjustment
                    TransactionLog.objects.create(
                        item=item,
                        order=order_item.order,
                        qty=-qty_accepted,
                        type=TransactionLog.TYPE_ADJUST,
                        user=request.user,
                        meta={
                            "reason": "return_scrap",
                            "return_no": return_obj.return_no,
                            "order_item_id": order_item.id,
                            "notes": notes,
                        },
                        notes=f"Return {return_obj.return_no} scrapped. {notes}",
                    )
                    
                    return_obj.status = Return.STATUS_SCRAPPED
                
                # Update return record
                return_obj.processed_at = timezone.now()
                return_obj.notes += f"\n[Processed] {notes}" if return_obj.notes else f"[Processed] {notes}"
                return_obj.save()
                
                context = {
                    "return_obj": return_obj,
                    "disposition": disposition,
                    "qty_accepted": qty_accepted,
                    "success": True,
                }
                return render(request, "inventory/partials/process_return_result.html", context)
        
        except Exception as e:
            context = {
                "success": False,
                "error_message": str(e),
            }
            return render(request, "inventory/partials/process_return_result.html", context)


# =============================
# Undo/Redo Views
# =============================

class UndoView(UserPassesTestMixin, View):
    """View to undo last N operations (manager-only)."""
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser
    
    def post(self, request, *args, **kwargs):
        """Perform undo operation."""
        count = int(request.POST.get("count", 1))
        
        try:
            results = perform_undo(user=request.user, count=count)
            
            context = {
                "results": results,
                "count": count,
                "success": True,
            }
            return render(request, "inventory/partials/undo_result.html", context)
            
        except Exception as e:
            context = {
                "success": False,
                "error_message": str(e),
            }
            return render(request, "inventory/partials/undo_result.html", context)


class RedoView(UserPassesTestMixin, View):
    """View to redo last N operations (manager-only)."""
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser
    
    def post(self, request, *args, **kwargs):
        """Perform redo operation."""
        count = int(request.POST.get("count", 1))
        
        try:
            results = perform_redo(user=request.user, count=count)
            
            context = {
                "results": results,
                "count": count,
                "success": True,
            }
            return render(request, "inventory/partials/redo_result.html", context)
            
        except Exception as e:
            context = {
                "success": False,
                "error_message": str(e),
            }
            return render(request, "inventory/partials/redo_result.html", context)


class UndoRedoHistoryView(UserPassesTestMixin, TemplateView):
    """View to display undo/redo stack history (manager-only)."""
    
    template_name = "inventory/undo_redo_history.html"
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        from .models import UndoStack, RedoStack
        
        # Get recent undo/redo operations
        context["undo_stack"] = UndoStack.objects.all()[:10]
        context["redo_stack"] = RedoStack.objects.all()[:10]
        
        return context


# =============================
# Notification Views
# =============================

class UnreadNotificationsView(LoginRequiredMixin, View):
    """View to return unread notification count and dropdown partial."""
    
    def get(self, request, *args, **kwargs):
        """Return unread count and notification list."""
        unread_count = get_unread_count(request.user)
        recent_notifications = get_recent_notifications(request.user, limit=5)
        
        context = {
            "unread_count": unread_count,
            "notifications": recent_notifications,
        }
        
        return render(request, "inventory/partials/notifications_dropdown.html", context)


class MarkNotificationReadView(LoginRequiredMixin, View):
    """View to mark a notification as read."""
    
    def post(self, request, *args, **kwargs):
        """Mark notification as read."""
        notification_id = kwargs.get("notification_id")
        success = mark_as_read(notification_id, request.user)
        
        return JsonResponse({"success": success})


class MarkAllNotificationsReadView(LoginRequiredMixin, View):
    """View to mark all notifications as read."""
    
    def post(self, request, *args, **kwargs):
        """Mark all notifications as read."""
        count = mark_all_as_read(request.user)
        
        return JsonResponse({"success": True, "count": count})


# =============================
# Webhook Endpoint
# =============================

@method_decorator(csrf_exempt, name="dispatch")
class WebhookReceiverView(View):
    """Endpoint for receiving external webhook events."""
    
    def post(self, request, *args, **kwargs):
        """Process incoming webhook."""
        # Validate signature
        if not validate_webhook_signature(request):
            logger.warning("Invalid webhook signature received")
            return JsonResponse({"error": "Invalid signature"}, status=401)
        
        try:
            # Parse payload
            payload = json.loads(request.body)
            event_type = payload.get("event_type")
            data = payload.get("data", {})
            
            # Process webhook based on event type
            if event_type == "external_order":
                self._handle_external_order(data)
            elif event_type == "inventory_sync":
                self._handle_inventory_sync(data)
            elif event_type == "notification":
                self._handle_notification(data)
            else:
                logger.warning(f"Unknown webhook event type: {event_type}")
            
            return JsonResponse({"status": "success", "event": event_type})
            
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception as e:
            logger.error(f"Webhook processing error: {e}")
            return JsonResponse({"error": str(e)}, status=500)
    
    def _handle_external_order(self, data):
        """Handle external order creation webhook."""
        order_no = data.get("order_no")
        customer = data.get("customer")
        
        # Create notification for managers
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        managers = User.objects.filter(is_staff=True)
        for manager in managers:
            notify(
                user=manager,
                message=f"External order received: {order_no} from {customer}",
                level="info",
                notification_type="order"
            )
        
        logger.info(f"External order webhook processed: {order_no}")
    
    def _handle_inventory_sync(self, data):
        """Handle inventory sync webhook."""
        item_sku = data.get("sku")
        qty_delta = data.get("qty_delta")
        
        # Create notification for inventory managers
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        managers = User.objects.filter(is_staff=True)
        for manager in managers:
            notify(
                user=manager,
                message=f"Inventory sync: {item_sku} adjusted by {qty_delta}",
                level="info",
                notification_type="inventory"
            )
        
        logger.info(f"Inventory sync webhook processed: {item_sku}")
    
    def _handle_notification(self, data):
        """Handle generic notification webhook."""
        message = data.get("message")
        level = data.get("level", "info")
        target_users = data.get("target_users", [])
        
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        if target_users:
            users = User.objects.filter(username__in=target_users)
        else:
            users = User.objects.filter(is_staff=True)
        
        for user in users:
            notify(
                user=user,
                message=message,
                level=level,
                notification_type="webhook"
            )
        
        logger.info(f"Notification webhook processed: {message[:50]}...")


# =============================
# Bulk Import Views
# =============================

class BulkImportView(UserPassesTestMixin, View):
    """View for bulk import of Items, Batches, or Orders."""
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser
    
    def get(self, request, *args, **kwargs):
        """Display upload form."""
        form = BulkImportForm()
        context = {"form": form}
        return render(request, "inventory/bulk_import.html", context)
    
    def post(self, request, *args, **kwargs):
        """Handle file upload and preview."""
        form = BulkImportForm(request.POST, request.FILES)
        
        if not form.is_valid():
            context = {"form": form}
            return render(request, "inventory/bulk_import.html", context)
        
        file = form.cleaned_data["file"]
        model_type = form.cleaned_data["model_type"]
        
        # Check if file should be queued for background processing
        file_size_threshold = 1 * 1024 * 1024  # 1MB
        
        if file.size > file_size_threshold:
            # Queue for background processing
            return self._queue_import(file, model_type, request.user)
        else:
            # Process immediately and show preview
            return self._preview_import(file, model_type)
    
    def _queue_import(self, file, model_type, user):
        """Queue import for background processing."""
        import os
        from django.conf import settings
        from django_q.tasks import async_task
        
        # Save file temporarily
        upload_dir = os.path.join(settings.MEDIA_ROOT, "imports")
        os.makedirs(upload_dir, exist_ok=True)
        
        file_path = os.path.join(upload_dir, f"{user.id}_{model_type}_{file.name}")
        
        with open(file_path, "wb+") as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        
        # Queue task
        task_id = async_task(
            "inventory.tasks.process_bulk_import",
            file_path,
            model_type,
            user.id,
        )
        
        notify(
            user=user,
            message=f"Import queued for processing: {file.name}",
            level="info",
            notification_type="import"
        )
        
        request = self.request
        context = {
            "queued": True,
            "task_id": task_id,
            "file_name": file.name,
        }
        
        return render(request, "inventory/partials/import_queued.html", context)
    
    def _preview_import(self, file, model_type):
        """Preview import data with validation."""
        import pandas as pd
        
        try:
            # Read file into dataframe
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
            
            # Validate data based on model type
            validation_results = self._validate_import_data(df, model_type)
            
            # Store data in session for commit
            preview_data = {
                "model_type": model_type,
                "data": df.to_dict('records'),
                "validation": validation_results,
            }
            
            # Store in session (simplified - in production use cache or temp storage)
            request = self.request
            request.session['import_preview'] = preview_data
            
            context = {
                "preview": True,
                "model_type": model_type,
                "rows": df.to_dict('records')[:100],  # Limit preview to 100 rows
                "total_rows": len(df),
                "validation_results": validation_results,
            }
            
            return render(request, "inventory/partials/import_preview.html", context)
            
        except Exception as e:
            logger.error(f"Import preview error: {e}")
            request = self.request
            context = {
                "error": str(e),
            }
            return render(request, "inventory/partials/import_error.html", context)
    
    def _validate_import_data(self, df, model_type):
        """Validate import data and return row-level errors."""
        import pandas as pd
        from datetime import datetime
        
        errors = []
        warnings = []
        
        if model_type == "item":
            # Validate Items
            for idx, row in df.iterrows():
                row_errors = []
                
                sku = str(row.get("sku", "")).strip()
                name = str(row.get("name", "")).strip()
                
                if not sku:
                    row_errors.append("Missing SKU")
                if not name:
                    row_errors.append("Missing name")
                
                if row_errors:
                    errors.append({
                        "row": idx + 2,  # +2 for header and 0-index
                        "errors": row_errors,
                    })
        
        elif model_type == "batch":
            # Validate Batches
            for idx, row in df.iterrows():
                row_errors = []
                
                item_sku = str(row.get("item_sku", "")).strip()
                lot_no = str(row.get("lot_no", "")).strip()
                
                if not item_sku:
                    row_errors.append("Missing item_sku")
                if not lot_no:
                    row_errors.append("Missing lot_no")
                
                # Validate received_qty
                try:
                    qty = float(row.get("received_qty", 0))
                    if qty <= 0:
                        row_errors.append("Invalid received_qty")
                except (ValueError, TypeError):
                    row_errors.append("Invalid received_qty format")
                
                # Validate expiry_date if present
                if pd.notna(row.get("expiry_date")):
                    try:
                        expiry = pd.to_datetime(row.get("expiry_date"))
                        if expiry.date() < datetime.now().date():
                            warnings.append({
                                "row": idx + 2,
                                "warning": "Expiry date in the past",
                            })
                    except Exception:
                        row_errors.append("Invalid expiry_date format")
                
                # Check if item exists
                if item_sku:
                    if not Item.objects.filter(sku=item_sku.upper()).exists():
                        row_errors.append(f"Item {item_sku} not found")
                
                if row_errors:
                    errors.append({
                        "row": idx + 2,
                        "errors": row_errors,
                    })
        
        elif model_type == "order":
            # Validate Orders
            for idx, row in df.iterrows():
                row_errors = []
                
                order_no = str(row.get("order_no", "")).strip()
                customer_name = str(row.get("customer_name", "")).strip()
                item_sku = str(row.get("item_sku", "")).strip()
                
                if not order_no:
                    row_errors.append("Missing order_no")
                if not customer_name:
                    row_errors.append("Missing customer_name")
                if not item_sku:
                    row_errors.append("Missing item_sku")
                
                # Validate qty_requested
                try:
                    qty = float(row.get("qty_requested", 0))
                    if qty <= 0:
                        row_errors.append("Invalid qty_requested")
                except (ValueError, TypeError):
                    row_errors.append("Invalid qty_requested format")
                
                # Check if item exists
                if item_sku:
                    if not Item.objects.filter(sku=item_sku.upper()).exists():
                        row_errors.append(f"Item {item_sku} not found")
                
                if row_errors:
                    errors.append({
                        "row": idx + 2,
                        "errors": row_errors,
                    })
        
        return {
            "errors": errors,
            "warnings": warnings,
            "valid_rows": len(df) - len(errors),
            "total_rows": len(df),
        }


class BulkImportCommitView(UserPassesTestMixin, View):
    """View to commit bulk import after preview."""
    
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser
    
    def post(self, request, *args, **kwargs):
        """Commit import data."""
        import pandas as pd
        
        # Retrieve preview data from session
        preview_data = request.session.get('import_preview')
        
        if not preview_data:
            context = {"error": "No import data found. Please upload again."}
            return render(request, "inventory/partials/import_error.html", context)
        
        model_type = preview_data["model_type"]
        data = preview_data["data"]
        df = pd.DataFrame(data)
        
        try:
            with transaction.atomic():
                if model_type == "item":
                    result = self._commit_items(df)
                elif model_type == "batch":
                    result = self._commit_batches(df)
                elif model_type == "order":
                    result = self._commit_orders(df)
                else:
                    raise ValueError(f"Unknown model type: {model_type}")
                
                # Clear session data
                del request.session['import_preview']
                
                notify(
                    user=request.user,
                    message=f"Import completed: {result['success']} {model_type}(s) imported",
                    level="success",
                    notification_type="import"
                )
                
                context = {
                    "success": True,
                    "result": result,
                    "model_type": model_type,
                }
                
                return render(request, "inventory/partials/import_success.html", context)
                
        except Exception as e:
            logger.error(f"Import commit error: {e}")
            context = {"error": str(e)}
            return render(request, "inventory/partials/import_error.html", context)
    
    def _commit_items(self, df):
        """Commit Item import."""
        success_count = 0
        
        for _, row in df.iterrows():
            sku = str(row.get("sku", "")).strip().upper()
            name = str(row.get("name", "")).strip()
            
            if not sku or not name:
                continue
            
            Item.objects.update_or_create(
                sku=sku,
                defaults={
                    "name": name,
                    "description": str(row.get("description", "")),
                    "unit": str(row.get("unit", "pcs")),
                    "reorder_threshold": Decimal(str(row.get("reorder_threshold", 0))),
                }
            )
            success_count += 1
        
        return {"success": success_count, "failed": len(df) - success_count}
    
    def _commit_batches(self, df):
        """Commit Batch import."""
        import pandas as pd
        
        success_count = 0
        
        for _, row in df.iterrows():
            item_sku = str(row.get("item_sku", "")).strip().upper()
            lot_no = str(row.get("lot_no", "")).strip()
            
            if not item_sku or not lot_no:
                continue
            
            try:
                item = Item.objects.get(sku=item_sku)
                received_qty = Decimal(str(row.get("received_qty", 0)))
                
                Batch.objects.create(
                    item=item,
                    lot_no=lot_no,
                    received_qty=received_qty,
                    available_qty=received_qty,
                    expiry_date=pd.to_datetime(row.get("expiry_date")).date() if pd.notna(row.get("expiry_date")) else None,
                    status=Batch.STATUS_AVAILABLE,
                )
                success_count += 1
            except Item.DoesNotExist:
                continue
        
        return {"success": success_count, "failed": len(df) - success_count}
    
    def _commit_orders(self, df):
        """Commit Order import."""
        success_count = 0
        
        # Group by order_no
        for order_no in df['order_no'].unique():
            order_rows = df[df['order_no'] == order_no]
            first_row = order_rows.iloc[0]
            
            try:
                # Create order
                order = Order.objects.create(
                    order_no=str(first_row.get("order_no", "")).strip(),
                    customer_name=str(first_row.get("customer_name", "")).strip(),
                    status=Order.STATUS_PENDING,
                )
                
                # Create order items
                for _, row in order_rows.iterrows():
                    item_sku = str(row.get("item_sku", "")).strip().upper()
                    item = Item.objects.get(sku=item_sku)
                    
                    OrderItem.objects.create(
                        order=order,
                        item=item,
                        qty_requested=Decimal(str(row.get("qty_requested", 0))),
                        status=OrderItem.STATUS_PENDING,
                    )
                
                success_count += 1
            except (Item.DoesNotExist, Exception):
                continue
        
        return {"success": success_count, "failed": len(df['order_no'].unique()) - success_count}


# =============================
# Export Views
# =============================

class ExportInventorySnapshotView(UserPassesTestMixin, View):
    """Export inventory snapshot with aggregated quantities."""
    
    def test_func(self):
        return self.request.user.is_authenticated
    
    def get(self, request, *args, **kwargs):
        """Export inventory snapshot."""
        format_type = request.GET.get("format", "csv")
        
        if format_type == "xlsx":
            return self._export_xlsx()
        else:
            return self._export_csv()
    
    def _export_csv(self):
        """Export as CSV with streaming response."""
        import csv
        from django.http import StreamingHttpResponse
        
        class Echo:
            """An object that implements just the write method of the file-like interface."""
            def write(self, value):
                return value
        
        def generate_rows():
            writer = csv.writer(Echo())
            # Header
            yield writer.writerow([
                'SKU', 'Name', 'Unit', 'Total Available Qty', 
                'Reorder Threshold', 'Status', 'Batches Count'
            ])
            
            # Data rows
            items = Item.objects.all().prefetch_related('batches')
            for item in items:
                total_qty = item.total_quantity()
                batches_count = item.batches.count()
                status = "Low Stock" if item.reorder_threshold and total_qty <= item.reorder_threshold else "OK"
                
                yield writer.writerow([
                    item.sku,
                    item.name,
                    item.unit,
                    total_qty,
                    item.reorder_threshold or 0,
                    status,
                    batches_count,
                ])
        
        response = StreamingHttpResponse(generate_rows(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="inventory_snapshot.csv"'
        return response
    
    def _export_xlsx(self):
        """Export as XLSX."""
        from openpyxl import Workbook
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Snapshot"
        
        # Header
        ws.append(['SKU', 'Name', 'Unit', 'Total Available Qty', 'Reorder Threshold', 'Status', 'Batches Count'])
        
        # Data
        items = Item.objects.all().prefetch_related('batches')
        for item in items:
            total_qty = item.total_quantity()
            batches_count = item.batches.count()
            status = "Low Stock" if item.reorder_threshold and total_qty <= item.reorder_threshold else "OK"
            
            ws.append([
                item.sku,
                item.name,
                item.unit,
                float(total_qty),
                float(item.reorder_threshold or 0),
                status,
                batches_count,
            ])
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="inventory_snapshot.xlsx"'
        wb.save(response)
        
        return response


class ExportBatchesView(UserPassesTestMixin, View):
    """Export all batches with full details."""
    
    def test_func(self):
        return self.request.user.is_authenticated
    
    def get(self, request, *args, **kwargs):
        """Export batches."""
        format_type = request.GET.get("format", "csv")
        
        if format_type == "xlsx":
            return self._export_xlsx()
        else:
            return self._export_csv()
    
    def _export_csv(self):
        """Export as CSV with streaming response."""
        import csv
        from django.http import StreamingHttpResponse
        
        class Echo:
            def write(self, value):
                return value
        
        def generate_rows():
            writer = csv.writer(Echo())
            # Header
            yield writer.writerow([
                'ID', 'Item SKU', 'Item Name', 'Lot No', 'Received Qty', 
                'Available Qty', 'Expiry Date', 'Status', 'Created At'
            ])
            
            # Data rows (streaming to handle large datasets)
            batches = Batch.objects.select_related('item').iterator(chunk_size=1000)
            for batch in batches:
                yield writer.writerow([
                    batch.id,
                    batch.item.sku,
                    batch.item.name,
                    batch.lot_no,
                    batch.received_qty,
                    batch.available_qty,
                    batch.expiry_date.isoformat() if batch.expiry_date else '',
                    batch.get_status_display(),
                    batch.created_at.isoformat(),
                ])
        
        response = StreamingHttpResponse(generate_rows(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="batches_export.csv"'
        return response
    
    def _export_xlsx(self):
        """Export as XLSX."""
        from openpyxl import Workbook
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Batches"
        
        # Header
        ws.append(['ID', 'Item SKU', 'Item Name', 'Lot No', 'Received Qty', 
                   'Available Qty', 'Expiry Date', 'Status', 'Created At'])
        
        # Data
        batches = Batch.objects.select_related('item').all()
        for batch in batches:
            ws.append([
                batch.id,
                batch.item.sku,
                batch.item.name,
                batch.lot_no,
                float(batch.received_qty),
                float(batch.available_qty),
                batch.expiry_date.isoformat() if batch.expiry_date else '',
                batch.get_status_display(),
                batch.created_at.isoformat(),
            ])
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="batches_export.xlsx"'
        wb.save(response)
        
        return response


class ExportTransactionLogView(UserPassesTestMixin, View):
    """Export transaction log with date filtering."""
    
    def test_func(self):
        return self.request.user.is_authenticated
    
    def get(self, request, *args, **kwargs):
        """Export transaction log."""
        format_type = request.GET.get("format", "csv")
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        
        if format_type == "xlsx":
            return self._export_xlsx(start_date, end_date)
        else:
            return self._export_csv(start_date, end_date)
    
    def _get_filtered_queryset(self, start_date, end_date):
        """Get filtered transaction log queryset."""
        from datetime import datetime
        
        qs = TransactionLog.objects.select_related(
            'user', 'batch__item', 'order_item__order', 'order_item__item'
        ).order_by('-timestamp')
        
        if start_date:
            try:
                start = datetime.fromisoformat(start_date)
                qs = qs.filter(timestamp__gte=start)
            except ValueError:
                pass
        
        if end_date:
            try:
                end = datetime.fromisoformat(end_date)
                qs = qs.filter(timestamp__lte=end)
            except ValueError:
                pass
        
        return qs
    
    def _export_csv(self, start_date, end_date):
        """Export as CSV with streaming response."""
        import csv
        from django.http import StreamingHttpResponse
        
        class Echo:
            def write(self, value):
                return value
        
        def generate_rows():
            writer = csv.writer(Echo())
            # Header
            yield writer.writerow([
                'ID', 'Timestamp', 'Transaction Type', 'User', 'Batch', 
                'Item SKU', 'Order No', 'Qty Change', 'Notes'
            ])
            
            # Data rows (streaming)
            transactions = self._get_filtered_queryset(start_date, end_date).iterator(chunk_size=1000)
            for txn in transactions:
                batch_info = f"{txn.batch.lot_no}" if txn.batch else ''
                item_sku = txn.item.sku if txn.item else ''
                order_no = txn.order.order_no if txn.order else ''
                user_name = txn.user.username if txn.user else 'System'
                notes = txn.meta.get('notes', '') if txn.meta else ''
                
                yield writer.writerow([
                    txn.id,
                    txn.timestamp.isoformat(),
                    txn.type,
                    user_name,
                    batch_info,
                    item_sku,
                    order_no,
                    txn.qty,
                    notes,
                ])
        
        response = StreamingHttpResponse(generate_rows(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="transaction_log.csv"'
        return response
    
    def _export_xlsx(self, start_date, end_date):
        """Export as XLSX."""
        from openpyxl import Workbook
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Transaction Log"
        
        # Header
        ws.append(['ID', 'Timestamp', 'Transaction Type', 'User', 'Batch', 
                   'Item SKU', 'Order No', 'Qty Change', 'Notes'])
        
        # Data
        transactions = self._get_filtered_queryset(start_date, end_date)
        for txn in transactions:
            batch_info = f"{txn.batch.lot_no}" if txn.batch else ''
            item_sku = txn.item.sku if txn.item else ''
            order_no = txn.order.order_no if txn.order else ''
            user_name = txn.user.username if txn.user else 'System'
            notes = txn.meta.get('notes', '') if txn.meta else ''
            
            ws.append([
                txn.id,
                txn.timestamp.isoformat(),
                txn.type,
                user_name,
                batch_info,
                item_sku,
                order_no,
                float(txn.qty),
                notes,
            ])
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="transaction_log.xlsx"'
        wb.save(response)
        
        return response


class ExportOrdersAllocationsView(UserPassesTestMixin, View):
    """Export orders with allocation details."""
    
    def test_func(self):
        return self.request.user.is_authenticated
    
    def get(self, request, *args, **kwargs):
        """Export orders and allocations."""
        format_type = request.GET.get("format", "csv")
        
        if format_type == "xlsx":
            return self._export_xlsx()
        else:
            return self._export_csv()
    
    def _export_csv(self):
        """Export as CSV with streaming response."""
        import csv
        from django.http import StreamingHttpResponse
        
        class Echo:
            def write(self, value):
                return value
        
        def generate_rows():
            writer = csv.writer(Echo())
            # Header
            yield writer.writerow([
                'Order No', 'Customer', 'Order Status', 'Item SKU', 'Item Name',
                'Qty Requested', 'Qty Picked', 'Item Status', 
                'Allocated Batch', 'Qty Allocated', 'Allocation Date'
            ])
            
            # Data rows (streaming)
            orders = Order.objects.prefetch_related(
                'items__item',
                'items__allocations__batch'
            ).iterator(chunk_size=500)
            
            for order in orders:
                for order_item in order.items.all():
                    allocations = order_item.allocations.all()
                    
                    if allocations:
                        for allocation in allocations:
                            yield writer.writerow([
                                order.order_no,
                                order.customer_name,
                                order.get_status_display(),
                                order_item.item.sku,
                                order_item.item.name,
                                order_item.qty_requested,
                                order_item.qty_picked or 0,
                                order_item.get_status_display(),
                                allocation.batch.lot_no,
                                allocation.qty_allocated,
                                allocation.allocated_at.isoformat(),
                            ])
                    else:
                        # Order item with no allocations
                        yield writer.writerow([
                            order.order_no,
                            order.customer_name,
                            order.get_status_display(),
                            order_item.item.sku,
                            order_item.item.name,
                            order_item.qty_requested,
                            order_item.qty_picked or 0,
                            order_item.get_status_display(),
                            '',
                            0,
                            '',
                        ])
        
        response = StreamingHttpResponse(generate_rows(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="orders_allocations.csv"'
        return response
    
    def _export_xlsx(self):
        """Export as XLSX."""
        from openpyxl import Workbook
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders & Allocations"
        
        # Header
        ws.append(['Order No', 'Customer', 'Order Status', 'Item SKU', 'Item Name',
                   'Qty Requested', 'Qty Picked', 'Item Status', 
                   'Allocated Batch', 'Qty Allocated', 'Allocation Date'])
        
        # Data
        orders = Order.objects.prefetch_related(
            'items__item',
            'items__allocations__batch'
        ).all()
        
        for order in orders:
            for order_item in order.items.all():
                allocations = order_item.allocations.all()
                
                if allocations:
                    for allocation in allocations:
                        ws.append([
                            order.order_no,
                            order.customer_name,
                            order.get_status_display(),
                            order_item.item.sku,
                            order_item.item.name,
                            float(order_item.qty_requested),
                            float(order_item.qty_picked or 0),
                            order_item.get_status_display(),
                            allocation.batch.lot_no,
                            float(allocation.qty_allocated),
                            allocation.allocated_at.isoformat(),
                        ])
                else:
                    ws.append([
                        order.order_no,
                        order.customer_name,
                        order.get_status_display(),
                        order_item.item.sku,
                        order_item.item.name,
                        float(order_item.qty_requested),
                        float(order_item.qty_picked or 0),
                        order_item.get_status_display(),
                        '',
                        0,
                        '',
                    ])
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="orders_allocations.xlsx"'
        wb.save(response)
        
        return response


class ExportMenuView(LoginRequiredMixin, TemplateView):
    """Menu view for export options."""
    
    template_name = "inventory/export_menu.html"


# =============================
# Graph Views (Cytoscape)
# =============================

class GraphDataView(LoginRequiredMixin, View):
    """Return inventory flow graph as Cytoscape-compatible JSON elements."""

    def get(self, request, *args, **kwargs):
        """Generate graph showing relationships between Items, Batches, Orders, and Shipments."""
        from django.db.models import Count, Sum
        
        elements = []
        node_ids = set()
        
        # Add Item nodes (colored blue)
        items = Item.objects.annotate(
            batch_count=Count('batches'),
            total_stock=Sum('batches__available_qty')
        ).filter(batch_count__gt=0)[:50]  # Limit to 50 for performance
        
        for item in items:
            node_id = f"item_{item.id}"
            node_ids.add(node_id)
            elements.append({
                "data": {
                    "id": node_id,
                    "label": f"{item.sku}\n({item.total_stock or 0} {item.unit})",
                    "type": "item",
                    "sku": item.sku,
                    "stock": float(item.total_stock or 0)
                },
                "classes": "item-node"
            })
        
        # Add Batch nodes (colored green) and connect to Items
        batches = Batch.objects.select_related('item').filter(
            available_qty__gt=0,
            status=Batch.STATUS_AVAILABLE
        )[:100]
        
        for batch in batches:
            batch_node_id = f"batch_{batch.id}"
            item_node_id = f"item_{batch.item.id}"
            
            if item_node_id in node_ids:
                node_ids.add(batch_node_id)
                elements.append({
                    "data": {
                        "id": batch_node_id,
                        "label": f"Lot: {batch.lot_no}\n{batch.available_qty} avail",
                        "type": "batch",
                        "lot_no": batch.lot_no,
                        "qty": float(batch.available_qty)
                    },
                    "classes": "batch-node"
                })
                
                # Edge: Item → Batch
                elements.append({
                    "data": {
                        "id": f"edge_item_batch_{batch.id}",
                        "source": item_node_id,
                        "target": batch_node_id,
                        "label": "has batch"
                    },
                    "classes": "item-batch-edge"
                })
        
        # Add recent Orders (colored orange) and connect to Items
        orders = Order.objects.filter(
            status__in=[Order.STATUS_NEW, Order.STATUS_ALLOCATED, Order.STATUS_PICKED]
        ).prefetch_related('items__item')[:30]
        
        for order in orders:
            order_node_id = f"order_{order.id}"
            node_ids.add(order_node_id)
            elements.append({
                "data": {
                    "id": order_node_id,
                    "label": f"{order.order_no}\n{order.get_status_display()}",
                    "type": "order",
                    "order_no": order.order_no,
                    "status": order.status
                },
                "classes": f"order-node status-{order.status}"
            })
            
            # Connect Order to Items it contains
            for order_item in order.items.all():
                item_node_id = f"item_{order_item.item.id}"
                if item_node_id in node_ids:
                    elements.append({
                        "data": {
                            "id": f"edge_order_item_{order.id}_{order_item.id}",
                            "source": order_node_id,
                            "target": item_node_id,
                            "label": f"needs {order_item.qty_requested}"
                        },
                        "classes": "order-item-edge"
                    })
        
        # Add Shipments (colored purple) and connect to Orders
        shipments = Shipment.objects.select_related('order').filter(
            status__in=[Shipment.STATUS_PENDING, Shipment.STATUS_IN_TRANSIT]
        )[:20]
        
        for shipment in shipments:
            shipment_node_id = f"shipment_{shipment.id}"
            order_node_id = f"order_{shipment.order.id}"
            
            if order_node_id in node_ids:
                node_ids.add(shipment_node_id)
                elements.append({
                    "data": {
                        "id": shipment_node_id,
                        "label": f"Ship: {shipment.tracking_no or 'Pending'}\n{shipment.get_status_display()}",
                        "type": "shipment",
                        "tracking": shipment.tracking_no or ''
                    },
                    "classes": "shipment-node"
                })
                
                # Edge: Order → Shipment
                elements.append({
                    "data": {
                        "id": f"edge_order_shipment_{shipment.id}",
                        "source": order_node_id,
                        "target": shipment_node_id,
                        "label": "shipped as"
                    },
                    "classes": "order-shipment-edge"
                })

        return JsonResponse({"elements": elements})


class GraphView(LoginRequiredMixin, TemplateView):
    template_name = "inventory/graph.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Inventory Flow Analysis'
        context['page_description'] = 'Visualize relationships between items, batches, orders, and shipments in real-time'
        return context


# =============================
# Stock Overview
# =============================

class StockOverviewView(LoginRequiredMixin, TemplateView):
    """Display stock overview with visualizations and inventory value."""
    template_name = "inventory/stock_overview.html"


class StockOverviewDataView(LoginRequiredMixin, View):
    """API endpoint for stock overview data with inventory valuation."""
    
    def get(self, request, *args, **kwargs):
        from django.db.models import Q, Sum, Count
        
        # Get all items with their stock calculations
        # Note: relations
        # Item -> Batch (related_name="batches")
        # Batch -> Allocation (related_name="allocations")
        # Allocation -> OrderItem -> Order (status is lower-case constants)
        items = Item.objects.annotate(
            available_stock=Sum(
                'batches__available_qty',
                filter=Q(batches__available_qty__gt=0),
                default=0
            ),
            allocated_stock=Sum(
                'batches__allocations__qty_allocated',
                filter=Q(
                    batches__allocations__order_item__order__status__in=[
                        Order.STATUS_NEW,
                        Order.STATUS_ALLOCATED,
                        Order.STATUS_PICKED,
                    ]
                ),
                default=0
            )
        ).order_by('-available_stock')
        
        # Calculate inventory data
        items_data = []
        total_value = Decimal('0.00')
        total_units = Decimal('0.00')
        low_stock_count = 0
        active_skus = 0
        
        for item in items:
            available = Decimal(str(item.available_stock or 0))
            allocated = Decimal(str(item.allocated_stock or 0))
            unit_price = item.price if item.price else Decimal('10.00')
            item_value = available * unit_price
            
            total_value += item_value
            total_units += available
            
            if available > 0:
                active_skus += 1
            
            if available <= (item.reorder_threshold or 0):
                low_stock_count += 1
            
            items_data.append({
                'id': item.id,
                'sku': item.sku,
                'name': item.name,
                'available_stock': float(available),
                'allocated_stock': float(allocated),
                'unit_price': float(unit_price),
                'total_value': float(item_value),
                'reorder_point': float(item.reorder_threshold or 0),
            })
        
        # Get batch count
        batch_count = Batch.objects.filter(available_qty__gt=0).count()
        
        summary = {
            'total_value': float(total_value),
            'total_skus': items.count(),
            'active_skus': active_skus,
            'total_units': float(total_units),
            'batch_count': batch_count,
            'low_stock_count': low_stock_count,
        }
        
        return JsonResponse({
            'summary': summary,
            'items': items_data
        })


# =============================
# Sandbox: Stack & Queue
# =============================

class SandboxStackView(LoginRequiredMixin, TemplateView):
    template_name = "inventory/sandbox_stack.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["can_apply"] = self.request.user.is_staff or self.request.user.is_superuser
        return ctx


class SandboxQueueView(LoginRequiredMixin, TemplateView):
    template_name = "inventory/sandbox_queue.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["can_apply"] = self.request.user.is_staff or self.request.user.is_superuser
        return ctx


class ApplySandboxOperationsView(UserPassesTestMixin, View):
    """Manager-only endpoint to apply sandbox operations to production.

    Expected JSON body:
      {
        "mode": "stack" | "queue",
        "operations": [{"sku": "SKU123", "qty": 5}, ...],
        "order_no": "OPTIONAL-ORDERNO"
      }
    """

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser

    def post(self, request, *args, **kwargs):
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except Exception:
            return HttpResponseBadRequest("Invalid JSON")

        mode = (payload.get("mode") or "").lower()
        operations = payload.get("operations") or []
        order_no = payload.get("order_no")

        if mode not in {"stack", "queue"}:
            return HttpResponseBadRequest("mode must be 'stack' or 'queue'")
        if not isinstance(operations, list) or not operations:
            return HttpResponseBadRequest("operations must be a non-empty list")

        # Normalize operations: list of dicts with sku, qty (Decimal)
        norm_ops = []
        from decimal import Decimal, InvalidOperation

        for op in operations:
            sku = (op.get("sku") or "").strip()
            qty = op.get("qty")
            if not sku:
                return HttpResponseBadRequest("operation missing sku")
            try:
                qty = Decimal(str(qty))
            except (InvalidOperation, TypeError):
                return HttpResponseBadRequest("invalid qty")
            if qty <= 0:
                return HttpResponseBadRequest("qty must be > 0")
            norm_ops.append({"sku": sku, "qty": qty})

        # Apply ordering semantics
        ops_in_order = list(norm_ops)
        if mode == "stack":
            ops_in_order = list(reversed(norm_ops))  # LIFO

        # Optional: link to order
        order = None
        if order_no:
            order = Order.objects.filter(order_no=order_no).first()
            if not order:
                return HttpResponseBadRequest("Order not found")

        # Pre-check availability per SKU
        from collections import defaultdict
        needed = defaultdict(Decimal)
        for op in ops_in_order:
            needed[op["sku"]] += op["qty"]

        for sku, total_needed in needed.items():
            item = Item.objects.filter(sku=sku).first()
            if not item:
                return HttpResponseBadRequest(f"Unknown SKU: {sku}")
            avail = Batch.objects.filter(item=item, status=Batch.STATUS_AVAILABLE, available_qty__gt=0).aggregate(total=Sum("available_qty")).get("total") or Decimal("0")
            if avail < total_needed:
                return HttpResponseBadRequest(f"Insufficient stock for {sku}: need {total_needed}, have {avail}")

        # Apply under a single transaction
        applied = []  # list of dicts: {batch_id, sku, qty}
        with transaction.atomic():
            for op in ops_in_order:
                sku = op["sku"]
                qty = op["qty"]
                item = Item.objects.select_for_update().get(sku=sku)
                qs = Batch.objects.select_for_update().filter(
                    item=item,
                    status=Batch.STATUS_AVAILABLE,
                    available_qty__gt=0,
                ).order_by("expiry_date", "id")

                remaining = qty
                for batch in qs:
                    if remaining <= 0:
                        break
                    can_take = min(batch.available_qty, remaining)
                    if can_take > 0:
                        batch.reserve(can_take)
                        TransactionLog.objects.create(
                            user=request.user,
                            type=TransactionLog.TYPE_RESERVE,
                            qty=-can_take,
                            item=item,
                            batch=batch,
                            order=order if order else None,
                            meta={"source": "sandbox_apply", "mode": mode, "note": "Applied from sandbox"},
                        )
                        applied.append({"batch_id": batch.id, "sku": sku, "qty": str(can_take)})
                        remaining -= can_take

                if remaining > 0:
                    # Should not happen due to pre-check; raise to rollback
                    raise ValueError(f"Insufficient stock after pre-check for {sku}")

            # Push single undo op with all batch adjustments
            push_undo_operation(
                op_name="reserve",
                metadata={"applied": applied, "order_id": order.id if order else None, "mode": mode},
            )

        return JsonResponse({"status": "ok", "applied": applied, "order_no": order_no or None})


# Order CRUD Views
class OrderListView(LoginRequiredMixin, ListView):
    """List all orders."""
    model = Order
    template_name = "inventory/order_list.html"
    context_object_name = "orders"
    paginate_by = 20

    def get_queryset(self):
        qs = Order.objects.prefetch_related('items__item').all()
        status_filter = self.request.GET.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)
        return qs


class OrderDetailView(LoginRequiredMixin, DetailView):
    """View order details with line items."""
    model = Order
    template_name = "inventory/order_detail.html"
    context_object_name = "order"
    pk_url_kwarg = "pk"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        order = self.get_object()
        context['order_items'] = order.items.select_related('item').all()
        context['allocations'] = Allocation.objects.filter(
            order_item__order=order
        ).select_related('batch', 'order_item__item')
        context['shipments'] = Shipment.objects.filter(order=order)
        return context


class OrderCreateView(LoginRequiredMixin, CreateView):
    """Create a new order with order items."""
    model = Order
    form_class = OrderForm
    template_name = "inventory/order_form.html"
    success_url = reverse_lazy('inventory:order-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['orderitem_formset'] = OrderItemInlineFormSet(
                self.request.POST,
                instance=self.object if self.object and self.object.pk else None
            )
        else:
            context['orderitem_formset'] = OrderItemInlineFormSet(
                instance=self.object if self.object and self.object.pk else None
            )
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        orderitem_formset = context['orderitem_formset']
        
        if orderitem_formset.is_valid():
            with transaction.atomic():
                self.object = form.save()
                orderitem_formset.instance = self.object
                orderitem_formset.save()
                
                notify(
                    user=self.request.user,
                    message=f"Order {self.object.order_no} created with {orderitem_formset.total_form_count()} items for {self.object.customer_name or 'N/A'}",
                    level="info",
                )
            
            from django.contrib import messages
            messages.success(self.request, f"Order {self.object.order_no} created successfully!")
            return super(CreateView, self).form_valid(form)
        else:
            return self.form_invalid(form)


class OrderUpdateView(LoginRequiredMixin, UpdateView):
    """Update order details and items."""
    model = Order
    form_class = OrderForm
    template_name = "inventory/order_form.html"
    success_url = reverse_lazy('inventory:order-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['orderitem_formset'] = OrderItemInlineFormSet(
                self.request.POST,
                instance=self.object
            )
        else:
            context['orderitem_formset'] = OrderItemInlineFormSet(
                instance=self.object
            )
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        orderitem_formset = context['orderitem_formset']
        
        if orderitem_formset.is_valid():
            with transaction.atomic():
                self.object = form.save()
                orderitem_formset.instance = self.object
                orderitem_formset.save()
            
            from django.contrib import messages
            messages.success(self.request, f"Order {self.object.order_no} updated successfully!")
            return super(UpdateView, self).form_valid(form)
        else:
            return self.form_invalid(form)


class OrderDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """Delete an order (manager only)."""
    model = Order
    template_name = "inventory/order_confirm_delete.html"
    success_url = reverse_lazy('inventory:order-list')

    def test_func(self):
        return self.request.user.is_staff


class OrderCancelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Cancel an order (manager only)."""

    def test_func(self):
        return self.request.user.is_staff

    def post(self, request, *args, **kwargs):
        order_id = kwargs.get("pk")
        
        try:
            order = Order.objects.get(pk=order_id)
            
            # Check if order can be cancelled
            if order.status in [Order.STATUS_CANCELLED, Order.STATUS_DELIVERED]:
                messages.warning(request, f"Order {order.order_no} cannot be cancelled (status: {order.get_status_display()}).")
                return redirect('inventory:order-detail', pk=order_id)
            
            # If order is allocated, deallocate first
            if order.status == Order.STATUS_ALLOCATED:
                # Deallocate inventory
                allocations = Allocation.objects.filter(order_item__order=order)
                
                with transaction.atomic():
                    for allocation in allocations:
                        # Return quantity to batch
                        Batch.objects.filter(pk=allocation.batch.pk).update(
                            available_qty=F('available_qty') + allocation.qty_allocated
                        )
                        
                        # Log the deallocation
                        TransactionLog.objects.create(
                            user=request.user,
                            type=TransactionLog.TYPE_UNRESERVE,
                            qty=allocation.qty_allocated,
                            item=allocation.order_item.item,
                            batch=allocation.batch,
                            order=order,
                            meta={'reason': 'order_cancelled', 'order_no': order.order_no}
                        )
                    
                    # Delete allocations
                    allocations.delete()
                    
                    # Reset order item allocations
                    OrderItem.objects.filter(order=order).update(qty_allocated=0)
            
            # Cancel the order
            order.status = Order.STATUS_CANCELLED
            order.save(update_fields=['status'])
            
            # Create notification
            Notification.objects.create(
                user=request.user,
                message=f"Order {order.order_no} has been cancelled.",
                level=Notification.LEVEL_WARNING,
            )
            
            messages.success(request, f"Order {order.order_no} has been successfully cancelled.")
            return redirect('inventory:order-detail', pk=order_id)
            
        except Order.DoesNotExist:
            messages.error(request, "Order not found.")
            return redirect('inventory:order-list')
        except Exception as e:
            messages.error(request, f"Failed to cancel order: {str(e)}")
            return redirect('inventory:order-detail', pk=order_id)


# ========================================
# Batch Order Processor Views (Queue + Stack Visualization)
# ========================================

class BatchProcessorView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """
    Batch order processor with Queue + Stack visualization.
    
    GET: Display the batch processor dashboard with queue stats.
    POST: Trigger batch processing of all NEW orders.
    """
    template_name = "inventory/batch_processor.html"

    def test_func(self):
        return self.request.user.is_staff

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get current queue statistics
        new_orders = Order.objects.filter(status='NEW').order_by('created_at')
        total_items = sum(
            order.orderitem_set.count() for order in new_orders
        )
        
        # Count available batches per item in queue
        items_needing_allocation = OrderItem.objects.filter(
            order__status='NEW'
        ).select_related('item')
        
        batch_availability = {}
        for order_item in items_needing_allocation:
            if order_item.item_id not in batch_availability:
                available_batches = Batch.objects.filter(
                    item=order_item.item,
                    available_qty__gt=0
                ).count()
                batch_availability[order_item.item_id] = available_batches
        
        context.update({
            'orders_in_queue': new_orders.count(),
            'total_items_in_queue': total_items,
            'unique_items': len(batch_availability),
            'total_batches_available': sum(batch_availability.values()),
            'new_orders': new_orders[:10],  # Preview first 10
        })
        
        return context

    def post(self, request, *args, **kwargs):
        """Trigger batch processing."""
        try:
            result = process_order_queue_batch(
                user=request.user,
                trace_enabled=False  # No trace for direct processing
            )
            
            messages.success(
                request,
                f"Processed {result['orders_processed']} orders. "
                f"Fully allocated: {result['fully_allocated']}, "
                f"Partially allocated: {result['partially_allocated']}, "
                f"Failed: {result['failed']}"
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Batch processing completed',
                'result': result
            })
            
        except Exception as e:
            messages.error(request, f"Batch processing failed: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class BatchProcessorTraceView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    API endpoint to execute batch processor and return full trace data.
    
    Returns JSON with step-by-step queue/stack operations for visualization.
    """

    def test_func(self):
        return self.request.user.is_staff

    def get(self, request, *args, **kwargs):
        """Run batch processor with tracing enabled."""
        try:
            result = process_order_queue_batch(
                user=request.user,
                trace_enabled=True
            )
            
            return JsonResponse({
                'success': True,
                'result': {
                    'orders_processed': result['orders_processed'],
                    'fully_allocated': result['fully_allocated'],
                    'partially_allocated': result['partially_allocated'],
                    'failed': result['failed'],
                },
                'trace': result.get('trace', {}),
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e),
                'traceback': traceback.format_exc()
            }, status=500)


# ========================================
# Warehouse Management Features
# ========================================

class LowStockAlertView(LoginRequiredMixin, TemplateView):
    """View for low stock alerts dashboard."""
    template_name = "inventory/low_stock_alert.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get all items with total quantities
        items_with_qty = Item.objects.annotate(
            total_qty=Sum('batches__available_qty', filter=Q(batches__status=Batch.STATUS_AVAILABLE))
        )
        
        # Critical: at or below reorder threshold
        critical_items = items_with_qty.filter(
            Q(total_qty__lte=F('reorder_threshold')) | Q(total_qty__isnull=True)
        ).order_by('total_qty')
        
        # Low: within 20% above reorder threshold
        low_items = []
        for item in items_with_qty:
            total = item.total_qty or 0
            threshold = item.reorder_threshold
            if threshold > 0 and total > threshold and total <= (threshold * Decimal('1.2')):
                low_items.append(item)
        
        # Healthy: above 20% margin
        healthy_count = items_with_qty.filter(
            total_qty__gt=F('reorder_threshold') * Decimal('1.2')
        ).count()
        
        context.update({
            'critical_items': critical_items,
            'low_items': low_items,
            'critical_count': critical_items.count(),
            'low_count': len(low_items),
            'healthy_count': healthy_count,
        })
        
        return context


class ExpiryTrackingView(LoginRequiredMixin, TemplateView):
    """View for batch expiry tracking dashboard."""
    template_name = "inventory/expiry_tracking.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        from datetime import timedelta
        today = timezone.now().date()
        
        # Get batches with expiry dates
        batches = Batch.objects.filter(
            status=Batch.STATUS_AVAILABLE,
            expiry_date__isnull=False,
            available_qty__gt=0
        ).select_related('item')
        
        # Categorize by expiry
        expired = []
        expiring_7 = []
        expiring_30 = []
        
        for batch in batches:
            days_diff = (batch.expiry_date - today).days
            batch.days_until_expiry = days_diff
            
            if days_diff < 0:
                expired.append(batch)
            elif days_diff <= 7:
                expiring_7.append(batch)
            elif days_diff <= 30:
                expiring_30.append(batch)
        
        # Count healthy batches
        healthy_count = Batch.objects.filter(
            status=Batch.STATUS_AVAILABLE,
            available_qty__gt=0
        ).filter(
            Q(expiry_date__isnull=True) | Q(expiry_date__gt=today + timedelta(days=30))
        ).count()
        
        context.update({
            'expired_batches': sorted(expired, key=lambda x: x.days_until_expiry),
            'expiring_7_batches': sorted(expiring_7, key=lambda x: x.days_until_expiry),
            'expiring_30_batches': sorted(expiring_30, key=lambda x: x.days_until_expiry),
            'expired_count': len(expired),
            'expiring_7_count': len(expiring_7),
            'expiring_30_count': len(expiring_30),
            'healthy_count': healthy_count,
        })
        
        return context


class InventoryMovementReportView(LoginRequiredMixin, TemplateView):
    """View for inventory movement report."""
    template_name = "inventory/movement_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get filter parameters
        item_sku = self.request.GET.get('item_sku')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        # Build query
        logs = TransactionLog.objects.select_related('item', 'batch', 'user', 'order').all()
        
        if item_sku:
            logs = logs.filter(item__sku=item_sku)
        
        if start_date:
            from datetime import datetime
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            logs = logs.filter(timestamp__gte=start_dt)
        
        if end_date:
            from datetime import datetime
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            logs = logs.filter(timestamp__lte=end_dt)
        
        logs = logs.order_by('-timestamp')[:200]  # Limit to 200 recent
        
        # Get all items for filter dropdown
        items = Item.objects.all().order_by('sku')
        
        context.update({
            'logs': logs,
            'items': items,
            'item_sku': item_sku,
            'start_date': start_date,
            'end_date': end_date,
        })
        
        return context


class LotSearchView(LoginRequiredMixin, TemplateView):
    """Lot/Batch search and detailed report."""
    template_name = "inventory/lot_search.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get search query
        lot_number = self.request.GET.get('lot_number', '').strip()
        start_date = self.request.GET.get('start_date', '').strip()
        end_date = self.request.GET.get('end_date', '').strip()
        
        batch = None
        transactions = []
        
        if lot_number:
            try:
                batch = Batch.objects.select_related('item').get(lot_no=lot_number)
                tx_qs = TransactionLog.objects.filter(batch=batch).select_related('item', 'user', 'order')
                # date range filters
                from datetime import datetime
                if start_date:
                    try:
                        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                        tx_qs = tx_qs.filter(timestamp__gte=start_dt)
                    except ValueError:
                        pass
                if end_date:
                    try:
                        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                        tx_qs = tx_qs.filter(timestamp__lte=end_dt)
                    except ValueError:
                        pass
                transactions = tx_qs.order_by('-timestamp')
            except Batch.DoesNotExist:
                messages.error(self.request, f'Lot number "{lot_number}" not found.')
        
        context.update({
            'lot_number': lot_number,
            'batch': batch,
            'transactions': transactions,
            'start_date': start_date,
            'end_date': end_date,
        })
        
        return context


class ItemSearchView(LoginRequiredMixin, TemplateView):
    """Item search with comprehensive stock details."""
    template_name = "inventory/item_search.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get search query
        item_sku = self.request.GET.get('item_sku', '').strip()
        start_date = self.request.GET.get('start_date', '').strip()
        end_date = self.request.GET.get('end_date', '').strip()
        
        item = None
        batches = []
        transactions = []
        allocations = []
        total_qty = 0
        suggestions = []
        
        if item_sku:
            try:
                item = Item.objects.get(sku__iexact=item_sku)
                
                # Get all batches for this item
                batches = Batch.objects.filter(item=item).order_by('-received_date')
                total_qty = batches.filter(status=Batch.STATUS_AVAILABLE).aggregate(
                    total=Sum('available_qty')
                )['total'] or 0
                
                # Get recent transactions
                tx_qs = TransactionLog.objects.filter(item=item).select_related('batch', 'user', 'order')
                from datetime import datetime
                if start_date:
                    try:
                        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                        tx_qs = tx_qs.filter(timestamp__gte=start_dt)
                    except ValueError:
                        pass
                if end_date:
                    try:
                        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                        tx_qs = tx_qs.filter(timestamp__lte=end_dt)
                    except ValueError:
                        pass
                transactions = tx_qs.order_by('-timestamp')[:100]
                
                # Get active allocations
                allocations = Allocation.objects.filter(
                    batch__item=item,
                    quantity__gt=0
                ).select_related('batch', 'order', 'order_item')
                
            except Item.DoesNotExist:
                # Provide suggestions by name/category
                suggestions = list(Item.objects.filter(name__icontains=item_sku).order_by('name')[:10])
                messages.error(self.request, f'Item SKU "{item_sku}" not found. Showing similar items by name.')
        
        context.update({
            'item_sku': item_sku,
            'item': item,
            'batches': batches,
            'transactions': transactions,
            'allocations': allocations,
            'total_qty': total_qty,
            'start_date': start_date,
            'end_date': end_date,
            'suggestions': suggestions,
        })
        
        return context


class CustomerSearchView(LoginRequiredMixin, TemplateView):
    """Customer order search and history."""
    template_name = "inventory/customer_search.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get search query
        customer_name = self.request.GET.get('customer_name', '').strip()
        start_date = self.request.GET.get('start_date', '').strip()
        end_date = self.request.GET.get('end_date', '').strip()
        
        orders = []
        order_count = 0
        total_value = 0.0
        order_totals = {}
        avg_value = 0.0
        
        if customer_name:
            qs = Order.objects.filter(customer_name__icontains=customer_name)
            # date range on created_at
            from datetime import datetime
            if start_date:
                try:
                    start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                    qs = qs.filter(created_at__date__gte=start_dt.date())
                except ValueError:
                    pass
            if end_date:
                try:
                    end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                    qs = qs.filter(created_at__date__lte=end_dt.date())
                except ValueError:
                    pass
            orders = qs.prefetch_related('items__item').order_by('-created_at')
            
            order_count = orders.count()
            
            # Calculate per-order and total values
            for order in orders:
                o_total = 0.0
                for oi in order.items.all():
                    try:
                        o_total += float(oi.item.price) * float(oi.quantity)
                    except Exception:
                        pass
                # attach for template access
                setattr(order, 'computed_total', o_total)
                order_totals[order.id] = o_total
                total_value += o_total
            if order_count:
                avg_value = total_value / order_count
        
        context.update({
            'customer_name': customer_name,
            'orders': orders,
            'order_count': order_count,
            'total_value': total_value,
            'avg_value': avg_value,
            'order_totals': order_totals,
            'start_date': start_date,
            'end_date': end_date,
        })
        
        return context


# ========================================
# Excel Export Views
# ========================================

class ExportLotReportView(LoginRequiredMixin, View):
    """Export lot/batch report to Excel."""
    
    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        lot_number = request.GET.get('lot_number', '').strip()
        
        if not lot_number:
            messages.error(request, 'Lot number is required.')
            return redirect('inventory:lot-search')
        
        try:
            batch = Batch.objects.select_related('item').get(lot_no=lot_number)
        except Batch.DoesNotExist:
            messages.error(request, f'Lot number "{lot_number}" not found.')
            return redirect('inventory:lot-search')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Lot {lot_number}"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Lot Information Section
        ws['A1'] = 'LOT/BATCH DETAILED REPORT'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = 'Lot Number:'
        ws['B3'] = batch.lot_no
        ws['A4'] = 'Item SKU:'
        ws['B4'] = batch.item.sku
        ws['A5'] = 'Item Name:'
        ws['B5'] = batch.item.name
        ws['A6'] = 'Received Date:'
        ws['B6'] = batch.received_date.strftime('%Y-%m-%d')
        ws['A7'] = 'Expiry Date:'
        ws['B7'] = batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else 'N/A'
        ws['A8'] = 'Initial Quantity:'
        ws['B8'] = float(batch.received_qty)
        ws['A9'] = 'Available Quantity:'
        ws['B9'] = float(batch.available_qty)
        ws['A10'] = 'Status:'
        ws['B10'] = batch.get_status_display()
        
        # Transaction History Section
        ws['A12'] = 'TRANSACTION HISTORY'
        ws['A12'].font = Font(size=14, bold=True)
        
        headers = ['Date/Time', 'Type', 'Quantity', 'User', 'Order ID', 'Notes']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=13, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Get transactions
        transactions = TransactionLog.objects.filter(
            batch=batch
        ).select_related('user', 'order').order_by('-timestamp')
        
        row = 14
        for txn in transactions:
            ws.cell(row=row, column=1, value=txn.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=2, value=txn.transaction_type)
            ws.cell(row=row, column=3, value=float(txn.quantity))
            ws.cell(row=row, column=4, value=txn.user.username if txn.user else 'System')
            ws.cell(row=row, column=5, value=f"#{txn.order.id}" if txn.order else 'N/A')
            ws.cell(row=row, column=6, value=txn.notes or '')
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 30
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="lot_{lot_number}_report.xlsx"'
        wb.save(response)
        
        return response


class ExportItemReportView(LoginRequiredMixin, View):
    """Export item stock report to Excel."""
    
    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        item_sku = request.GET.get('item_sku', '').strip()
        
        if not item_sku:
            messages.error(request, 'Item SKU is required.')
            return redirect('inventory:item-search')
        
        try:
            item = Item.objects.get(sku__iexact=item_sku)
        except Item.DoesNotExist:
            messages.error(request, f'Item SKU "{item_sku}" not found.')
            return redirect('inventory:item-search')
        
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: Item Summary
        ws1 = wb.active
        ws1.title = "Item Summary"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        ws1['A1'] = 'ITEM STOCK DETAILED REPORT'
        ws1['A1'].font = Font(size=16, bold=True)
        ws1.merge_cells('A1:D1')
        
        ws1['A3'] = 'SKU:'
        ws1['B3'] = item.sku
        ws1['A4'] = 'Name:'
        ws1['B4'] = item.name
        ws1['A5'] = 'Category:'
        ws1['B5'] = item.category
        ws1['A6'] = 'Description:'
        ws1['B6'] = item.description
        ws1['A7'] = 'Unit:'
        ws1['B7'] = item.unit
        ws1['A8'] = 'Price:'
        ws1['B8'] = float(item.price)
        ws1['A9'] = 'Reorder Threshold:'
        ws1['B9'] = float(item.reorder_threshold)
        
        # Get total quantity
        total_qty = Batch.objects.filter(
            item=item, status=Batch.STATUS_AVAILABLE
        ).aggregate(total=Sum('available_qty'))['total'] or 0
        
        ws1['A10'] = 'Total Available:'
        ws1['B10'] = float(total_qty)
        
        # Sheet 2: Batches
        ws2 = wb.create_sheet("Batches")
        
        headers = ['Lot Number', 'Received Date', 'Expiry Date', 'Initial Qty', 'Available Qty', 'Status']
        for col_num, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        batches = Batch.objects.filter(item=item).order_by('-received_date')
        row = 2
        for batch in batches:
            ws2.cell(row=row, column=1, value=batch.lot_no)
            ws2.cell(row=row, column=2, value=batch.received_date.strftime('%Y-%m-%d'))
            ws2.cell(row=row, column=3, value=batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else 'N/A')
            ws2.cell(row=row, column=4, value=float(batch.initial_qty))
            ws2.cell(row=row, column=5, value=float(batch.available_qty))
            ws2.cell(row=row, column=6, value=batch.get_status_display())
            row += 1
        
        # Sheet 3: Transactions
        ws3 = wb.create_sheet("Transactions")
        
        headers = ['Date/Time', 'Type', 'Lot Number', 'Quantity', 'User', 'Order ID']
        for col_num, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        transactions = TransactionLog.objects.filter(
            item=item
        ).select_related('batch', 'user', 'order').order_by('-timestamp')[:100]
        
        row = 2
        for txn in transactions:
            ws3.cell(row=row, column=1, value=txn.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
            ws3.cell(row=row, column=2, value=txn.transaction_type)
            ws3.cell(row=row, column=3, value=txn.batch.lot_no if txn.batch else 'N/A')
            ws3.cell(row=row, column=4, value=float(txn.quantity))
            ws3.cell(row=row, column=5, value=txn.user.username if txn.user else 'System')
            ws3.cell(row=row, column=6, value=f"#{txn.order.id}" if txn.order else 'N/A')
            row += 1
        
        # Adjust column widths
        for ws in [ws1, ws2, ws3]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="item_{item.sku}_report.xlsx"'
        wb.save(response)
        
        return response


class ExportCustomerReportView(LoginRequiredMixin, View):
    """Export customer order history to Excel."""
    
    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        customer_name = request.GET.get('customer_name', '').strip()
        
        if not customer_name:
            messages.error(request, 'Customer name is required.')
            return redirect('inventory:customer-search')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Customer Orders"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        ws['A1'] = f'CUSTOMER ORDER HISTORY: {customer_name}'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:G1')
        
        # Headers
        headers = ['Order ID', 'Created Date', 'Status', 'Priority', 'Total Items', 'Total Value', 'Notes']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Get orders
        orders = Order.objects.filter(
            customer_name__icontains=customer_name
        ).prefetch_related('items__item').order_by('-created_at')
        
        row = 4
        total_value_all = 0
        
        for order in orders:
            total_items = order.items.count()
            order_value = sum(
                float(oi.item.price) * float(oi.quantity) for oi in order.items.all()
            )
            total_value_all += order_value
            
            ws.cell(row=row, column=1, value=f"#{order.id}")
            ws.cell(row=row, column=2, value=order.created_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=3, value=order.get_status_display())
            ws.cell(row=row, column=4, value=order.get_priority_display())
            ws.cell(row=row, column=5, value=total_items)
            ws.cell(row=row, column=6, value=order_value)
            ws.cell(row=row, column=7, value=order.notes or '')
            row += 1
        
        # Summary
        ws.cell(row=row + 1, column=1, value='TOTAL ORDERS:')
        ws.cell(row=row + 1, column=1).font = Font(bold=True)
        ws.cell(row=row + 1, column=2, value=orders.count())
        
        ws.cell(row=row + 2, column=1, value='TOTAL VALUE:')
        ws.cell(row=row + 2, column=1).font = Font(bold=True)
        ws.cell(row=row + 2, column=2, value=total_value_all)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 30
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="customer_{customer_name}_orders.xlsx"'
        wb.save(response)
        
        return response



# ========================================
# CSV Export Views
# ========================================

class ExportLotReportCSVView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        import csv
        from django.http import HttpResponse
        from datetime import datetime
        lot_number = request.GET.get('lot_number', '').strip()
        if not lot_number:
            messages.error(request, 'Lot number is required.')
            return redirect('inventory:lot-search')
        try:
            batch = Batch.objects.select_related('item').get(lot_no=lot_number)
        except Batch.DoesNotExist:
            messages.error(request, f'Lot number "{lot_number}" not found.')
            return redirect('inventory:lot-search')
        # Optional date filtering
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        start_dt = None
        end_dt = None
        try:
            if start_date:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            if end_date:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        except Exception:
            start_dt = end_dt = None
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="lot_{lot_number}_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['LOT/BATCH DETAILED REPORT'])
        writer.writerow([])
        writer.writerow(['Lot Number', batch.lot_no])
        writer.writerow(['Item SKU', batch.item.sku])
        writer.writerow(['Item Name', batch.item.name])
        writer.writerow(['Received Date', batch.received_date.strftime('%Y-%m-%d')])
        writer.writerow(['Expiry Date', batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else 'N/A'])
        writer.writerow(['Initial Qty', float(batch.received_qty)])
        writer.writerow(['Available Qty', float(batch.available_qty)])
        writer.writerow(['Status', batch.get_status_display()])
        writer.writerow([])
        writer.writerow(['Date/Time', 'Type', 'Quantity', 'User', 'Order ID', 'Notes'])
        txns = TransactionLog.objects.filter(batch=batch)
        if start_dt:
            txns = txns.filter(timestamp__date__gte=start_dt.date())
        if end_dt:
            txns = txns.filter(timestamp__date__lte=end_dt.date())
        txns = txns.select_related('user','order').order_by('-timestamp')
        for txn in txns:
            writer.writerow([
                txn.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                txn.transaction_type,
                float(txn.quantity),
                txn.user.username if txn.user else 'System',
                f"#{txn.order.id}" if txn.order else 'N/A',
                txn.notes or ''
            ])
        return response


class ExportItemReportCSVView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        import csv
        from django.http import HttpResponse
        from datetime import datetime
        item_sku = request.GET.get('item_sku', '').strip()
        if not item_sku:
            messages.error(request, 'Item SKU is required.')
            return redirect('inventory:item-search')
        try:
            item = Item.objects.get(sku__iexact=item_sku)
        except Item.DoesNotExist:
            messages.error(request, f'Item SKU "{item_sku}" not found.')
            return redirect('inventory:item-search')
        # Optional date filtering
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        start_dt = None
        end_dt = None
        try:
            if start_date:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            if end_date:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        except Exception:
            start_dt = end_dt = None
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="item_{item.sku}_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['ITEM STOCK DETAILED REPORT'])
        writer.writerow([])
        writer.writerow(['SKU', item.sku])
        writer.writerow(['Name', item.name])
        writer.writerow(['Category', item.category])
        writer.writerow(['Description', item.description])
        writer.writerow(['Unit', item.unit])
        writer.writerow(['Price', float(item.price)])
        total_qty = Batch.objects.filter(item=item, status=Batch.STATUS_AVAILABLE).aggregate(total=Sum('available_qty'))['total'] or 0
        writer.writerow(['Total Available', float(total_qty)])
        writer.writerow([])
        writer.writerow(['Batches'])
        writer.writerow(['Lot Number', 'Received Date', 'Expiry Date', 'Initial Qty', 'Available Qty', 'Status'])
        for b in Batch.objects.filter(item=item).order_by('-received_date'):
            writer.writerow([
                b.lot_no,
                b.received_date.strftime('%Y-%m-%d'),
                b.expiry_date.strftime('%Y-%m-%d') if b.expiry_date else 'N/A',
                float(b.received_qty),
                float(b.available_qty),
                b.get_status_display(),
            ])
        writer.writerow([])
        writer.writerow(['Transactions (Last 100)'])
        writer.writerow(['Date/Time', 'Type', 'Lot Number', 'Quantity', 'User', 'Order ID'])
        txns = TransactionLog.objects.filter(item=item)
        if start_dt:
            txns = txns.filter(timestamp__date__gte=start_dt.date())
        if end_dt:
            txns = txns.filter(timestamp__date__lte=end_dt.date())
        txns = txns.select_related('batch','user','order').order_by('-timestamp')[:100]
        for txn in txns:
            writer.writerow([
                txn.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                txn.transaction_type,
                txn.batch.lot_no if txn.batch else 'N/A',
                float(txn.quantity),
                txn.user.username if txn.user else 'System',
                f"#{txn.order.id}" if txn.order else 'N/A',
            ])
        return response


class ExportCustomerReportCSVView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        import csv
        from django.http import HttpResponse
        from datetime import datetime
        customer_name = request.GET.get('customer_name', '').strip()
        if not customer_name:
            messages.error(request, 'Customer name is required.')
            return redirect('inventory:customer-search')
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="customer_{customer_name}_orders.csv"'
        writer = csv.writer(response)
        writer.writerow([f'CUSTOMER ORDER HISTORY: {customer_name}'])
        writer.writerow([])
        writer.writerow(['Order ID', 'Created Date', 'Status', 'Priority', 'Total Items', 'Total Value', 'Notes'])
        orders = Order.objects.filter(customer_name__icontains=customer_name)
        # Optional date filters on orders
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        try:
            if start_date:
                sd = datetime.strptime(start_date, '%Y-%m-%d').date()
                orders = orders.filter(created_at__date__gte=sd)
            if end_date:
                ed = datetime.strptime(end_date, '%Y-%m-%d').date()
                orders = orders.filter(created_at__date__lte=ed)
        except Exception:
            pass
        orders = orders.prefetch_related('items__item').order_by('-created_at')
        total_value_all = 0
        for order in orders:
            total_items = order.items.count()
            order_value = sum(float(oi.item.price) * float(oi.quantity) for oi in order.items.all())
            total_value_all += order_value
            writer.writerow([
                f"#{order.id}",
                order.created_at.strftime('%Y-%m-%d %H:%M'),
                order.get_status_display(),
                order.get_priority_display(),
                total_items,
                order_value,
                order.notes or ''
            ])
        writer.writerow([])
        writer.writerow(['TOTAL ORDERS', orders.count()])
        writer.writerow(['TOTAL VALUE', total_value_all])
        return response


